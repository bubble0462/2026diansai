from __future__ import annotations

import argparse
import os
import queue
import re
import shutil
import sys
import tempfile
import threading
import time
import winreg
from statistics import mean, stdev
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable

import serial
from serial.tools import list_ports
from openpyxl import load_workbook


BAUD_RATE = 115200


def windows_desktop() -> Path:
    try:
        key_path = r"Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders"
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path) as key:
            value, _ = winreg.QueryValueEx(key, "Desktop")
        return Path(os.path.expandvars(value))
    except OSError:
        return Path.home() / "Desktop"


DEFAULT_WORKBOOK = windows_desktop() / "拟合.xlsx"
SHEET_NAME = "标定数据"
TERMINATOR = b"\xff\xff\xff"
READY = b"\xfe\xff\xff\xff"
FINISH = b"\xfd\xff\xff\xff"
VALUE_RE = re.compile(rb'^(tf1|tupp|turms)\.txt="([^"]*)"$')
NUMBER_RE = re.compile(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)")
ADDT_RE = re.compile(rb"^addt\s+\d+,\d+,(\d+)$")
CAL_RE = re.compile(
    rb"^CAL,F=([-+]?\d+(?:\.\d+)?),"
    rb"VPP=([-+]?\d+(?:\.\d+)?),"
    rb"RMS=([-+]?\d+(?:\.\d+)?)$"
)


@dataclass(frozen=True)
class Measurement:
    sequence: int
    frequency_khz: float
    vpp_mv: float
    rms_mv: float
    received_at: float


@dataclass(frozen=True)
class Target:
    row: int
    frequency_khz: float
    source_vpp_mv: float


def parse_number(text: str) -> float | None:
    match = NUMBER_RE.search(text)
    return float(match.group(0)) if match else None


class TjcStreamParser:
    def __init__(self) -> None:
        self.buffer = bytearray()
        self.binary_remaining = 0
        self.values: dict[str, float] = {}
        self.sequence = 0

    def feed(self, data: bytes) -> tuple[list[Measurement], list[bytes]]:
        self.buffer.extend(data)
        measurements: list[Measurement] = []
        responses: list[bytes] = []

        while self.buffer:
            if self.binary_remaining:
                count = min(self.binary_remaining, len(self.buffer))
                del self.buffer[:count]
                self.binary_remaining -= count
                if self.binary_remaining == 0:
                    responses.append(FINISH)
                continue

            end = self.buffer.find(TERMINATOR)
            if end < 0:
                if len(self.buffer) > 8192:
                    del self.buffer[:-256]
                break

            packet = bytes(self.buffer[:end])
            del self.buffer[: end + len(TERMINATOR)]
            if not packet:
                continue

            addt = ADDT_RE.match(packet)
            if addt:
                self.binary_remaining = int(addt.group(1))
                responses.append(READY)
                continue

            value = VALUE_RE.match(packet)
            if not value:
                continue

            name = value.group(1).decode("ascii")
            parsed = parse_number(value.group(2).decode("ascii", errors="ignore"))
            if parsed is None:
                self.values.pop(name, None)
                continue

            if name == "tf1":
                self.values = {"tf1": parsed}
            else:
                self.values[name] = parsed

            if all(key in self.values for key in ("tf1", "tupp", "turms")):
                self.sequence += 1
                measurements.append(
                    Measurement(
                        sequence=self.sequence,
                        frequency_khz=self.values["tf1"],
                        vpp_mv=self.values["tupp"],
                        rms_mv=self.values["turms"],
                        received_at=time.time(),
                    )
                )
                self.values = {}

        return measurements, responses


class CalibrationStreamParser:
    def __init__(self) -> None:
        self.buffer = bytearray()
        self.sequence = 0

    def feed(self, data: bytes) -> tuple[list[Measurement], list[bytes]]:
        self.buffer.extend(data)
        measurements: list[Measurement] = []
        while True:
            end = self.buffer.find(b"\n")
            if end < 0:
                if len(self.buffer) > 1024:
                    del self.buffer[:-128]
                break
            line = bytes(self.buffer[:end]).strip()
            del self.buffer[: end + 1]
            match = CAL_RE.match(line)
            if not match:
                continue
            self.sequence += 1
            measurements.append(
                Measurement(
                    sequence=self.sequence,
                    frequency_khz=float(match.group(1)),
                    vpp_mv=float(match.group(2)),
                    rms_mv=float(match.group(3)),
                    received_at=time.time(),
                )
            )
        return measurements, []


class SerialWorker(threading.Thread):
    def __init__(
        self,
        port: str,
        baud: int,
        output: queue.Queue[tuple[str, object]],
    ) -> None:
        super().__init__(daemon=True)
        self.port = port
        self.baud = baud
        self.output = output
        self.stop_event = threading.Event()
        self.serial: serial.Serial | None = None

    def stop(self) -> None:
        self.stop_event.set()
        if self.serial and self.serial.is_open:
            try:
                self.serial.cancel_read()
            except (AttributeError, OSError, serial.SerialException):
                pass

    def run(self) -> None:
        parser = CalibrationStreamParser()
        try:
            self.serial = serial.Serial(
                self.port,
                self.baud,
                bytesize=serial.EIGHTBITS,
                parity=serial.PARITY_NONE,
                stopbits=serial.STOPBITS_ONE,
                timeout=0.1,
                write_timeout=0.5,
            )
            self.output.put(("connected", self.port))
            while not self.stop_event.is_set():
                data = self.serial.read(self.serial.in_waiting or 1)
                if not data:
                    continue
                measurements, responses = parser.feed(data)
                for response in responses:
                    self.serial.write(response)
                for measurement in measurements:
                    self.output.put(("measurement", measurement))
        except Exception as exc:
            self.output.put(("error", str(exc)))
        finally:
            if self.serial and self.serial.is_open:
                self.serial.close()
            self.output.put(("disconnected", self.port))


def workbook_targets(path: Path) -> list[Target]:
    book = load_workbook(path, read_only=True, data_only=False)
    try:
        if SHEET_NAME not in book.sheetnames:
            raise ValueError(f"工作簿中找不到“{SHEET_NAME}”工作表")
        sheet = book[SHEET_NAME]
        targets = []
        for row in range(2, sheet.max_row + 1):
            frequency = sheet.cell(row, 1).value
            source_vpp = sheet.cell(row, 2).value
            if isinstance(frequency, (int, float)) and isinstance(source_vpp, (int, float)):
                targets.append(Target(row, float(frequency), float(source_vpp)))
        if not targets:
            raise ValueError("标定数据中没有有效的频率/幅度行")
        # Calibration order is independent of the physical Excel row order:
        # hold Vpp constant, sweep 10 kHz -> 500 kHz, then advance Vpp.
        targets.sort(
            key=lambda target: (
                target.source_vpp_mv,
                target.frequency_khz,
                target.row,
            )
        )
        return targets
    finally:
        book.close()


def _write_with_running_excel(
    path: Path,
    target: Target,
    sample_index: int,
    measurement: Measurement,
) -> bool:
    try:
        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            return False

        workbook = None
        wanted = os.path.normcase(os.path.abspath(path))
        for candidate in excel.Workbooks:
            if os.path.normcase(os.path.abspath(candidate.FullName)) == wanted:
                workbook = candidate
                break
        if workbook is None:
            return False

        sheet = workbook.Worksheets(SHEET_NAME)
        vpp_col = 4 + sample_index
        rms_col = 10 + sample_index
        sheet.Cells(target.row, vpp_col).Value = measurement.vpp_mv
        sheet.Cells(target.row, rms_col).Value = measurement.rms_mv
        sheet.Cells(target.row, 18).Value = (
            f"串口实测频率 {measurement.frequency_khz:.3f} kHz；"
            f"{datetime.now():%Y-%m-%d %H:%M:%S}"
        )
        excel.Calculate()
        workbook.Save()
        return True
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _write_batch_with_running_excel(
    path: Path,
    target: Target,
    start_index: int,
    measurements: list[Measurement],
) -> bool:
    try:
        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            return False

        workbook = None
        wanted = os.path.normcase(os.path.abspath(path))
        for candidate in excel.Workbooks:
            if os.path.normcase(os.path.abspath(candidate.FullName)) == wanted:
                workbook = candidate
                break
        if workbook is None:
            return False

        sheet = workbook.Worksheets(SHEET_NAME)
        for offset, measurement in enumerate(measurements):
            sample_index = start_index + offset
            sheet.Cells(target.row, 4 + sample_index).Value = measurement.vpp_mv
            sheet.Cells(target.row, 10 + sample_index).Value = measurement.rms_mv
        frequencies = [measurement.frequency_khz for measurement in measurements]
        sheet.Cells(target.row, 18).Value = (
            f"串口批量采集 {len(measurements)} 帧；"
            f"频率 {min(frequencies):.3f}～{max(frequencies):.3f} kHz；"
            f"{datetime.now():%Y-%m-%d %H:%M:%S}"
        )
        excel.Calculate()
        workbook.Save()
        return True
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def write_measurements(
    path: Path,
    target: Target,
    start_index: int,
    measurements: list[Measurement],
) -> str:
    if not measurements:
        raise ValueError("没有可写入的测量数据。")
    if start_index < 0 or (start_index + len(measurements)) > 3:
        raise ValueError("批量测量数量超出 Excel 的3次记录列。")
    if _write_batch_with_running_excel(path, target, start_index, measurements):
        return "已一次性写入当前打开的 Excel"

    book = load_workbook(path)
    try:
        sheet = book[SHEET_NAME]
        for offset, measurement in enumerate(measurements):
            sample_index = start_index + offset
            sheet.cell(target.row, 4 + sample_index, measurement.vpp_mv)
            sheet.cell(target.row, 10 + sample_index, measurement.rms_mv)
        frequencies = [measurement.frequency_khz for measurement in measurements]
        sheet.cell(
            target.row,
            18,
            f"串口批量采集 {len(measurements)} 帧；"
            f"频率 {min(frequencies):.3f}～{max(frequencies):.3f} kHz；"
            f"{datetime.now():%Y-%m-%d %H:%M:%S}",
        )
        book.calculation.fullCalcOnLoad = True
        book.calculation.forceFullCalc = True
        book.calculation.calcMode = "auto"

        handle, temp_name = tempfile.mkstemp(
            prefix=f".{path.stem}_",
            suffix=path.suffix,
            dir=path.parent,
        )
        os.close(handle)
        temp_path = Path(temp_name)
        try:
            book.save(temp_path)
            shutil.copystat(path, temp_path)
            os.replace(temp_path, path)
        finally:
            temp_path.unlink(missing_ok=True)
    finally:
        book.close()
    return "已一次性写入工作簿"


def write_measurement(
    path: Path,
    target: Target,
    sample_index: int,
    measurement: Measurement,
) -> str:
    return write_measurements(path, target, sample_index, [measurement])


def used_samples(path: Path, target: Target) -> int:
    book = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet = book[SHEET_NAME]
        count = 0
        for sample_index in range(3):
            vpp = sheet.cell(target.row, 4 + sample_index).value
            rms = sheet.cell(target.row, 10 + sample_index).value
            if vpp not in (None, "") and rms not in (None, ""):
                count += 1
            else:
                break
        return count
    finally:
        book.close()


class CalibrationApp:
    def __init__(self, root, workbook: Path) -> None:
        import tkinter as tk
        from tkinter import filedialog, messagebox, ttk

        self.tk = tk
        self.ttk = ttk
        self.filedialog = filedialog
        self.messagebox = messagebox
        self.root = root
        self.root.title("STM32 串口标定采集")
        self.root.geometry("820x650")
        self.root.minsize(760, 600)

        self.events: queue.Queue[tuple[str, object]] = queue.Queue()
        self.worker: SerialWorker | None = None
        self.targets: list[Target] = []
        self.target_index = 0
        self.latest: Measurement | None = None
        self.capture_after_sequence = -1
        self.capture_deadline = 0.0
        self.waiting_capture = False
        self.capture_target: Target | None = None
        self.capture_target_index = -1
        self.capture_start_index = 0
        self.capture_needed = 0
        self.capture_measurements: list[Measurement] = []

        self.workbook_var = tk.StringVar(value=str(workbook))
        self.port_var = tk.StringVar()
        self.status_var = tk.StringVar(value="未连接")
        self.latest_var = tk.StringVar(value="等待串口数据")
        self.batch_result_var = tk.StringVar(value="点击一次可自动连续采集3帧")
        self.source_frequency_var = tk.StringVar()
        self.source_vpp_var = tk.StringVar()
        self.progress_var = tk.StringVar(value="")
        self.samples_var = tk.IntVar(value=3)
        self.auto_next_var = tk.BooleanVar(value=True)

        self._build_ui()
        self.reload_workbook()
        self.refresh_ports()
        self.root.after(80, self.poll_events)
        self.root.protocol("WM_DELETE_WINDOW", self.close)

    def _build_ui(self) -> None:
        tk = self.tk
        ttk = self.ttk
        outer = ttk.Frame(self.root, padding=14)
        outer.pack(fill="both", expand=True)

        file_box = ttk.LabelFrame(outer, text="标定工作簿", padding=10)
        file_box.pack(fill="x")
        ttk.Entry(file_box, textvariable=self.workbook_var).pack(
            side="left", fill="x", expand=True
        )
        ttk.Button(file_box, text="浏览", command=self.browse_workbook).pack(
            side="left", padx=(8, 0)
        )
        ttk.Button(file_box, text="重新加载", command=self.reload_workbook).pack(
            side="left", padx=(8, 0)
        )
        ttk.Button(file_box, text="打开 Excel", command=self.open_workbook).pack(
            side="left", padx=(8, 0)
        )

        serial_box = ttk.LabelFrame(outer, text="串口", padding=10)
        serial_box.pack(fill="x", pady=(10, 0))
        ttk.Label(serial_box, text="端口").pack(side="left")
        self.port_combo = ttk.Combobox(
            serial_box, textvariable=self.port_var, width=28, state="readonly"
        )
        self.port_combo.pack(side="left", padx=(8, 6))
        ttk.Button(serial_box, text="刷新", command=self.refresh_ports).pack(side="left")
        self.connect_button = ttk.Button(
            serial_box, text="连接", command=self.toggle_connection
        )
        self.connect_button.pack(side="left", padx=(8, 0))
        ttk.Label(serial_box, text=f"{BAUD_RATE} / 8N1").pack(side="left", padx=14)
        ttk.Label(serial_box, textvariable=self.status_var).pack(side="right")

        target_box = ttk.LabelFrame(outer, text="当前信号源设置", padding=12)
        target_box.pack(fill="x", pady=(10, 0))
        settings = ttk.Frame(target_box)
        settings.pack()
        ttk.Label(
            settings,
            text="频率",
            font=("Microsoft YaHei UI", 13, "bold"),
        ).pack(side="left")
        self.frequency_entry = ttk.Combobox(
            settings,
            textvariable=self.source_frequency_var,
            width=12,
            font=("Microsoft YaHei UI", 14),
        )
        self.frequency_entry.pack(side="left", padx=(8, 5))
        ttk.Label(settings, text="kHz").pack(side="left")
        ttk.Label(
            settings,
            text="峰峰值",
            font=("Microsoft YaHei UI", 13, "bold"),
        ).pack(side="left", padx=(28, 0))
        self.vpp_entry = ttk.Combobox(
            settings,
            textvariable=self.source_vpp_var,
            width=12,
            font=("Microsoft YaHei UI", 14),
        )
        self.vpp_entry.pack(side="left", padx=(8, 5))
        ttk.Label(settings, text="mVpp").pack(side="left")
        ttk.Button(settings, text="匹配表格行", command=self.match_input_target).pack(
            side="left", padx=(20, 0)
        )
        ttk.Label(target_box, textvariable=self.progress_var).pack(pady=(4, 0))
        navigation = ttk.Frame(target_box)
        navigation.pack(pady=(10, 0))
        ttk.Button(navigation, text="◀ 上一点", command=self.previous_target).pack(
            side="left"
        )
        ttk.Button(navigation, text="下一点 ▶", command=self.next_target).pack(
            side="left", padx=(10, 0)
        )
        ttk.Label(navigation, text="每点记录").pack(side="left", padx=(24, 6))
        ttk.Combobox(
            navigation,
            textvariable=self.samples_var,
            values=(1, 2, 3),
            width=3,
            state="readonly",
        ).pack(side="left")
        ttk.Label(navigation, text="次").pack(side="left", padx=(4, 0))
        ttk.Checkbutton(
            navigation,
            text="完成后自动下一频率（扫完后换幅度）",
            variable=self.auto_next_var,
        ).pack(side="left", padx=(20, 0))

        reading_box = ttk.LabelFrame(outer, text="串口最新值", padding=12)
        reading_box.pack(fill="x", pady=(10, 0))
        ttk.Label(
            reading_box,
            textvariable=self.latest_var,
            font=("Consolas", 16, "bold"),
        ).pack()
        ttk.Label(
            reading_box,
            textvariable=self.batch_result_var,
            font=("Microsoft YaHei UI", 11),
            foreground="#1F4E78",
        ).pack(pady=(6, 0))
        self.capture_button = ttk.Button(
            reading_box,
            text="自动采集并一次性写入当前行",
            command=self.request_capture,
        )
        self.capture_button.pack(pady=(12, 0), ipadx=24, ipady=8)

        log_box = ttk.LabelFrame(outer, text="操作记录", padding=8)
        log_box.pack(fill="both", expand=True, pady=(10, 0))
        self.log_text = tk.Text(
            log_box,
            height=10,
            state="disabled",
            font=("Consolas", 10),
            wrap="word",
        )
        scroll = ttk.Scrollbar(log_box, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scroll.set)
        self.log_text.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        ttk.Label(
            outer,
            text="独立标定串口：USART2，PA2=TX、PA3=RX；串口屏继续使用 PA9/PA10。",
            foreground="#9C5700",
        ).pack(anchor="w", pady=(8, 0))

    def log(self, message: str) -> None:
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{datetime.now():%H:%M:%S}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def workbook_path(self) -> Path:
        return Path(self.workbook_var.get()).expanduser().resolve()

    def browse_workbook(self) -> None:
        filename = self.filedialog.askopenfilename(
            title="选择标定工作簿",
            filetypes=[("Excel 工作簿", "*.xlsx")],
            initialdir=str(self.workbook_path().parent),
        )
        if filename:
            self.workbook_var.set(filename)
            self.reload_workbook()

    def reload_workbook(self) -> None:
        try:
            path = self.workbook_path()
            self.targets = workbook_targets(path)
            self.target_index = min(self.target_index, len(self.targets) - 1)
            frequencies = sorted({target.frequency_khz for target in self.targets})
            amplitudes = sorted({target.source_vpp_mv for target in self.targets})
            self.frequency_entry["values"] = [f"{value:g}" for value in frequencies]
            self.vpp_entry["values"] = [f"{value:g}" for value in amplitudes]
            if not self.source_frequency_var.get():
                self.source_frequency_var.set(
                    f"{self.targets[self.target_index].frequency_khz:g}"
                )
            if not self.source_vpp_var.get():
                self.source_vpp_var.set(
                    f"{self.targets[self.target_index].source_vpp_mv:g}"
                )
            self.match_input_target(show_error=False)
            self.update_target_display()
            self.log(f"已加载 {path}，共 {len(self.targets)} 个标定点")
        except Exception as exc:
            self.targets = []
            self.progress_var.set("工作簿加载失败")
            self.messagebox.showerror("工作簿错误", str(exc))

    def open_workbook(self) -> None:
        try:
            os.startfile(self.workbook_path())
        except Exception as exc:
            self.messagebox.showerror("无法打开 Excel", str(exc))

    def refresh_ports(self) -> None:
        ports = list(list_ports.comports())
        labels = [f"{p.device} — {p.description}" for p in ports]
        self.port_combo["values"] = labels
        current = self.port_var.get()
        if labels and current not in labels:
            self.port_var.set(labels[0])
        elif not labels:
            self.port_var.set("")
        self.log(f"发现 {len(labels)} 个串口")

    def selected_port(self) -> str:
        return self.port_var.get().split(" — ", 1)[0].strip()

    def toggle_connection(self) -> None:
        if self.worker and self.worker.is_alive():
            self.disconnect()
            return
        port = self.selected_port()
        if not port:
            self.messagebox.showwarning("未选择串口", "请先选择串口。")
            return
        self.worker = SerialWorker(port, BAUD_RATE, self.events)
        self.worker.start()
        self.status_var.set("正在连接…")
        self.connect_button.configure(text="断开")
        self.log(f"正在连接 {port}")

    def disconnect(self) -> None:
        if self.worker:
            self.worker.stop()
        self.reset_capture()
        self.status_var.set("正在断开…")

    def reset_capture(self) -> None:
        self.waiting_capture = False
        self.capture_target = None
        self.capture_target_index = -1
        self.capture_start_index = 0
        self.capture_needed = 0
        self.capture_measurements = []
        self.capture_button.configure(
            state="normal", text="自动采集并一次性写入当前行"
        )

    def previous_target(self) -> None:
        if self.targets:
            self.target_index = max(0, self.target_index - 1)
            self.set_input_from_target(self.targets[self.target_index])
            self.update_target_display()

    def next_target(self) -> None:
        if self.targets:
            self.target_index = min(len(self.targets) - 1, self.target_index + 1)
            self.set_input_from_target(self.targets[self.target_index])
            self.update_target_display()

    def set_input_from_target(self, target: Target) -> None:
        self.source_frequency_var.set(f"{target.frequency_khz:g}")
        self.source_vpp_var.set(f"{target.source_vpp_mv:g}")

    def input_target(self) -> tuple[int, Target]:
        try:
            frequency = float(self.source_frequency_var.get().strip())
            source_vpp = float(self.source_vpp_var.get().strip())
        except ValueError as exc:
            raise ValueError("请输入有效的数字频率和峰峰值。") from exc

        for index, target in enumerate(self.targets):
            if (
                abs(target.frequency_khz - frequency) < 1e-6
                and abs(target.source_vpp_mv - source_vpp) < 1e-6
            ):
                return index, target
        raise ValueError(
            f"Excel 中没有 {frequency:g} kHz / {source_vpp:g} mVpp 对应行。"
            "请使用下拉列表中的标定值，或先在表格中增加该点。"
        )

    def match_input_target(self, show_error: bool = True) -> bool:
        try:
            index, _ = self.input_target()
            self.target_index = index
            self.update_target_display()
            return True
        except Exception as exc:
            self.progress_var.set(str(exc))
            if show_error:
                self.messagebox.showwarning("找不到对应行", str(exc))
            return False

    def update_target_display(self) -> None:
        if not self.targets:
            return
        target = self.targets[self.target_index]
        amplitude_targets = [
            item
            for item in self.targets
            if item.source_vpp_mv == target.source_vpp_mv
        ]
        frequency_index = amplitude_targets.index(target) + 1
        amplitudes = sorted({item.source_vpp_mv for item in self.targets})
        amplitude_index = amplitudes.index(target.source_vpp_mv) + 1
        try:
            used = used_samples(self.workbook_path(), target)
        except Exception:
            used = 0
        self.progress_var.set(
            f"已匹配 Excel 第 {target.row} 行；"
            f"幅度 {amplitude_index}/{len(amplitudes)}，"
            f"该幅度频率 {frequency_index}/{len(amplitude_targets)}；"
            f"当前已记录 {used}/3 次"
        )

    def request_capture(self) -> None:
        if not (self.worker and self.worker.is_alive()):
            self.messagebox.showwarning("串口未连接", "请先连接串口。")
            return
        if not self.targets:
            self.messagebox.showwarning("没有标定点", "请先加载工作簿。")
            return
        try:
            target_index, target = self.input_target()
        except Exception as exc:
            self.messagebox.showwarning("无法读取", str(exc))
            return
        self.target_index = target_index
        self.capture_target = target
        self.capture_target_index = target_index
        wanted = max(1, min(3, int(self.samples_var.get())))
        try:
            existing = used_samples(self.workbook_path(), target)
        except Exception as exc:
            self.reset_capture()
            self.messagebox.showerror("无法读取工作簿", str(exc))
            return
        if existing >= wanted:
            self.reset_capture()
            self.messagebox.showwarning(
                "当前点已完成",
                f"当前行已经记录 {existing} 次，达到设定的 {wanted} 次。",
            )
            return
        self.capture_start_index = existing
        self.capture_needed = wanted - existing
        self.capture_measurements = []
        self.update_target_display()
        self.capture_after_sequence = self.latest.sequence if self.latest else -1
        self.capture_deadline = time.monotonic() + 10.0
        self.waiting_capture = True
        self.capture_button.configure(
            state="disabled", text=f"正在采集 0/{self.capture_needed} 帧…"
        )
        self.batch_result_var.set(
            f"正在连续采集：0/{self.capture_needed} 帧，完成后一次性写入"
        )
        self.log(
            f"开始批量采集：当前已有 {existing} 帧，"
            f"本次自动补采 {self.capture_needed} 帧"
        )

    def handle_capture(self, measurement: Measurement) -> None:
        target = self.capture_target
        if target is None:
            self.reset_capture()
            return
        self.target_index = self.capture_target_index
        difference = abs(measurement.frequency_khz - target.frequency_khz)
        if difference > 2.0:
            self.reset_capture()
            self.messagebox.showwarning(
                "频率不匹配",
                f"当前行要求 {target.frequency_khz:g} kHz，"
                f"串口读到 {measurement.frequency_khz:.3f} kHz。\n"
                "未写入，请检查信号源设置。",
            )
            self.log("频率不匹配，已拒绝写入")
            return

        try:
            wanted = max(1, min(3, int(self.samples_var.get())))
            self.capture_measurements.append(measurement)
            collected = len(self.capture_measurements)
            self.capture_after_sequence = measurement.sequence
            self.capture_button.configure(
                text=f"正在采集 {collected}/{self.capture_needed} 帧…"
            )
            self.batch_result_var.set(
                f"正在连续采集：{collected}/{self.capture_needed} 帧"
            )
            self.log(
                f"已缓存第 {collected}/{self.capture_needed} 帧："
                f"F={measurement.frequency_khz:.3f} kHz，"
                f"Vpp={measurement.vpp_mv:.2f} mV，"
                f"RMS={measurement.rms_mv:.2f} mV"
            )
            if collected < self.capture_needed:
                return

            current_used = used_samples(self.workbook_path(), target)
            if current_used != self.capture_start_index:
                raise ValueError("采集期间 Excel 当前行发生变化，为防止覆盖，本批数据未写入。")
            result = write_measurements(
                self.workbook_path(),
                target,
                self.capture_start_index,
                self.capture_measurements,
            )
            vpps = [item.vpp_mv for item in self.capture_measurements]
            rmss = [item.rms_mv for item in self.capture_measurements]
            vpp_std = stdev(vpps) if len(vpps) >= 2 else None
            rms_std = stdev(rmss) if len(rmss) >= 2 else None
            vpp_std_text = f"{vpp_std:.3f}" if vpp_std is not None else "—"
            rms_std_text = f"{rms_std:.3f}" if rms_std is not None else "—"
            self.batch_result_var.set(
                f"本批：Vpp均值 {mean(vpps):.3f} mV，标准差 {vpp_std_text} mV；"
                f"RMS均值 {mean(rmss):.3f} mV，标准差 {rms_std_text} mV"
            )
            self.log(
                f"{result}：第 {target.row} 行，从第 "
                f"{self.capture_start_index + 1} 次开始，共 {collected} 帧"
            )
            self.log(self.batch_result_var.get())
            completed = self.capture_start_index + collected >= wanted
            if (
                completed
                and self.auto_next_var.get()
                and self.target_index < len(self.targets) - 1
            ):
                previous_amplitude = target.source_vpp_mv
                self.target_index += 1
                next_target = self.targets[self.target_index]
                self.set_input_from_target(next_target)
                if next_target.source_vpp_mv == previous_amplitude:
                    self.log(
                        f"当前点完成：幅度保持 {next_target.source_vpp_mv:g} mVpp，"
                        f"请将频率调到 {next_target.frequency_khz:g} kHz"
                    )
                else:
                    self.log(
                        f"当前幅度的全部频率已完成：请将幅度调到 "
                        f"{next_target.source_vpp_mv:g} mVpp，并将频率调回 "
                        f"{next_target.frequency_khz:g} kHz"
                    )
            self.update_target_display()
        except Exception as exc:
            self.messagebox.showerror("写入失败", str(exc))
            self.log(f"写入失败：{exc}")
        finally:
            if len(self.capture_measurements) >= self.capture_needed:
                self.reset_capture()

    def poll_events(self) -> None:
        try:
            while True:
                kind, payload = self.events.get_nowait()
                if kind == "connected":
                    self.status_var.set(f"已连接 {payload}")
                    self.log(f"串口 {payload} 已连接")
                elif kind == "measurement":
                    measurement = payload
                    assert isinstance(measurement, Measurement)
                    self.latest = measurement
                    self.latest_var.set(
                        f"F = {measurement.frequency_khz:9.3f} kHz    "
                        f"Vpp = {measurement.vpp_mv:8.2f} mV    "
                        f"RMS = {measurement.rms_mv:8.2f} mV"
                    )
                    if (
                        self.waiting_capture
                        and measurement.sequence > self.capture_after_sequence
                    ):
                        self.handle_capture(measurement)
                elif kind == "error":
                    self.status_var.set("串口错误")
                    self.log(f"串口错误：{payload}")
                    self.messagebox.showerror("串口错误", str(payload))
                elif kind == "disconnected":
                    self.status_var.set("未连接")
                    self.connect_button.configure(text="连接")
                    self.reset_capture()
                    self.log(f"串口 {payload} 已断开")
        except queue.Empty:
            pass

        if self.waiting_capture and time.monotonic() > self.capture_deadline:
            collected = len(self.capture_measurements)
            needed = self.capture_needed
            self.reset_capture()
            self.log(
                f"读取超时：10秒内仅收到 {collected}/{needed} 帧有效 CAL 数据"
            )
            self.messagebox.showwarning(
                "读取超时",
                f"10秒内仅收到 {collected}/{needed} 帧有效 CAL 数据。\n"
                "请检查 TX/RX/GND、端口和 115200 波特率。",
            )
        self.root.after(80, self.poll_events)

    def close(self) -> None:
        if self.worker:
            self.worker.stop()
        self.root.destroy()


def self_test() -> None:
    parser = CalibrationStreamParser()
    measurements, responses = parser.feed(
        b"noise\r\nCAL,F=100."
        b"001,VPP=247.50,RMS=87.51\r\nCAL,INVALID\r\n"
    )
    assert len(measurements) == 1
    assert measurements[0].frequency_khz == 100.001
    assert measurements[0].vpp_mv == 247.5
    assert measurements[0].rms_mv == 87.51
    assert not responses

    tjc = TjcStreamParser()
    stream = (
        b'tf1.txt="100.001 kHz"\xff\xff\xff'
        b'tupp.txt="247.50 mV"\xff\xff\xff'
        b'turms.txt="87.51 mV"\xff\xff\xff'
        b"addt 1,0,4\xff\xff\xff"
        b"\x01\x02\x03\x04"
    )
    old_measurements, old_responses = tjc.feed(stream[:31])
    assert not old_measurements
    more, more_responses = tjc.feed(stream[31:])
    old_measurements += more
    old_responses += more_responses
    assert len(old_measurements) == 1
    assert old_responses == [READY, FINISH]
    print("self-test passed")


def main() -> None:
    parser = argparse.ArgumentParser(description="STM32 串口标定采集工具")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()
    if args.self_test:
        self_test()
        return

    import tkinter as tk

    root = tk.Tk()
    CalibrationApp(root, args.workbook)
    root.mainloop()


if __name__ == "__main__":
    main()
